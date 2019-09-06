import PyPDF2 as p
from openpyxl import load_workbook
import os
import datetime
import shutil

def sheet_creator(part_list, unit, calendars):
    if len(part_list) == 0:
        return
    excel_template = "ECG Summary Spreadsheet Template.xlsx"
    wb = load_workbook(excel_template)
    ws = wb.active
    row = 6
    wo_comp = WO_company(part_list[0][1], unit, calendars) #(date of first part, unit#, list of calendars)
    for person in part_list:
        ws['B' + str(row)] = person[3]
        ws['C' + str(row)] = person[0]
        ws['D' + str(row)] = person[2]
        ws['E' + str(row)] = person[4]
        row += 1
    ws['D1'] = unit
    ws['F1'] = 'Date: ' + str(part_list[0][1])
    ws['C1'] = wo_comp[0] #Sets company's name
    ws['E3'] = wo_comp[1] #Sets WO
    ws['C3'] = wo_comp[2] #Sets Location name
    wb.save(unit + '_' + person[1] + '.xlsx')
    wb.close()
    print(1)
    return
    
def maker(file_list):
    parts = []
    for x in file_list:
        if '.pdf' in x:
            file = open(x, 'rb')
            pdfReader = p.PdfFileReader(file)
            pdfObj = pdfReader.getPage(0)
            text = pdfObj.extractText()
            file.close()
            
            name = finder(text, ("Patient Name", "Patient unique number"))
            name = name[14:]                                #JUST the patient's name. Strip takes off the end characters, so can't use it!
            dob = finder(text, ("DOB", "10mm/mV")).strip("DOB: ")     #Patient DOB
            date = (text[text.find("Run:") - 11:text.find("Run:") - 1])     #Date of test
            part_number = (finder(text, ("Patient unique number:", "Age:")).strip("Patient unique number: "))    #Patient Participant Number
            ecg_number = 1
            
            parts.append( [name, date, dob, part_number, ecg_number] )
            print("Working...")
            
    parts = sorted(parts)        #Makes our list alphabetized
    parts = number_checker(parts)
    
    return parts

def number_checker(parts):
    for x in parts:
        count = parts.count(x)
        if count > 1:
            for i in range(count - 1):
                parts.pop(parts.index(x) + 1)
            x[4] = count
               
    return parts

def finder(pdf, words):
    start = pdf.find(words[0])
    end = pdf.find(words[1])
    
    return pdf[start:end]

def files_in_folder(temp):
    file_list = []
    
    for (dirpath, dirnames, filenames) in os.walk(temp):
        file_list.extend(filenames)
        break
    
    return file_list

def folder_finder(temp):
    folder_list = []
    
    for (dirpath, dirnames, filenames) in os.walk(temp):
        folder_list.extend(dirnames)
        break

    return folder_list

def file_deleter(file_list):
    for file in file_list:
        os.remove(file)
    return

def unit_date_sorter(part_list):
    dates = dict()
    if len(part_list) > 0:
        for person in part_list:
            if len(dates) == 0:
                dates[person[1]] = [person]
            elif len(dates.keys()) > 0:
                if person[1] in dates.keys():
                    dates[person[1]].append(person)
                else:
                    dates[person[1]] = [person] 
    return dates

def ecg_cover_maker(cover_info):
    cover_template = 'ECG Fax Cover Template.xlsx'
    wb = load_workbook(cover_template)
    ws = wb.active
    ws['C1'] = datetime.date.today()
    total_ecg = 0
    for thing in cover_info:
        total_ecg += int(thing[2])
    row = 5
    for line in cover_info:
        ws['A' + str(row)] = line[0]
        ws['B' + str(row)] = line[1]
        ws['C' + str(row)] = line[2]
        row += 1
    ws['A' + str(row)] = 'Total Number of Tracings'
    ws['C' + str(row)] = total_ecg
    wb.save('ECG_Fax_Cover' + str(datetime.date.today()) + '.xlsx')
    wb.close()
    return

def ecg_count(parts):
    ecg_count = 0
    for participant in parts:
        ecg_count += int(participant[4])
    return ecg_count

def WO_company(date, unit, calendars):
    file = open("Cal_key.txt")
    keys = file.read().split(',')
    
    if ('3' in unit) or ('4' in unit) or ('8' in unit) or ('9' in unit):
        for name in calendars[0].get_sheet_names():
            if str(unit) in name:
                break
        ws = calendars[0].get_sheet_by_name(name)
        for day in keys:
            if (date[:4] + "-" + date[5:7] + "-" + date[8:]) in day:
                row = day[:(day.find(':'))]
        company = ws['D' + str(row)].value
        WO = ws['G' + str(row)].value
        try:
            location = (ws['E' + str(row)].value + " " + ws['F' + str(row)].value)
        except TypeError:
            print("No location given! Uh oh!")
            location = ""
        return [company, WO, location]
        
    elif ('2' in unit) or ('1' in unit) or ('10' in unit):
        for name in calendars[1].get_sheet_names():
            if str(unit) in name:
                break
        ws = calendars[1].get_sheet_by_name(name)
        for day in keys:
            if (date[:4] + "-" + date[5:7] + "-" + date[8:]) in day:
                row = day[:(day.find(':'))]
        company = ws['D' + str(row)].value
        WO = ws['I' + str(row)].value
        try:
            location = (ws['E' + str(row)].value + " " + ws['F' + str(row)].value)
        except TypeError:
            print("No location given! Uh oh!")
            location = ""
        return [company, WO, location]

def calendar_loader():
    location = '\\\\zebi\\departments\\Operations\\Shared\\MMP\\Units\\Unit_Cal\\2017'
    os.chdir(location)
    wb1 = load_workbook(filename='2017Cal.xlsx', read_only=True)
    wb2 = load_workbook(filename='Unit Schedule_2017_BB.xlsx', read_only=True)
    print("Calendars loaded")
    return [wb1, wb2]

	
def main():
    cover_info = []
    calendars = calendar_loader()
    temp = 'C:\\Users\\adrian.ridder\\Desktop\\temp'
    os.chdir(temp)
    
    for unit in folder_finder(temp):
        os.chdir(temp + '\\' + unit)
        part_list = maker(files_in_folder(os.getcwd()))
        sorted_dates = unit_date_sorter(part_list)                      #File handling is "done" after this point
        for date in sorted_dates.keys():
            cover_info.append([unit, date, ecg_count(sorted_dates[date])]) #appends list w/ format: unit, date, ecg_count
            try:
                os.makedirs(date)
            except FileExistsError:
                pass
            for file in files_in_folder(os.getcwd()):
                if str(date) in file:
                    shutil.move(os.getcwd() + "\\" + file, os.getcwd() + "\\" + date + "\\" + file)	
        os.chdir(temp)
        for date in sorted_dates.keys():
            sheet_creator(sorted_dates[date], unit, calendars)
    ecg_cover_maker(cover_info)
    print("Done")
    
if __name__ == '__main__':
    main()

