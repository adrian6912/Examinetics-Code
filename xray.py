from openpyxl import load_workbook
import os
import datetime as dt
import dicom as d

def sheet_creator(part_list):
    if len(part_list) == 0:
        return
    excel_template = "Mobile Xray Example Sheet.xlsx"
    wb = load_workbook(excel_template)
    ws = wb.active
    row = 11
    for person in part_list:
        ws['D' + str(row)] = person[3]
        ws['E' + str(row)] = person[0]
        ws['F' + str(row)] = person[2]
        ws['G' + str(row)] = "Y"
        ws['H' + str(row)] = person[4]
        row += 1
    wb.save(person[1] + '.xlsx')
    wb.close()
    print(1)
    return

def unit_date_sorter(part_list):
    dates = dict()
    if len(part_list) > 0:
        for person in part_list:
            if len(dates) == 0:
                dates[person[1]] = [person]
            elif len(dates.keys()) > 0:
                if person[1] in dates.keys():
                    dates[person[1]].append(person)
                else:
                    dates[person[1]] = [person] 
    return dates
    
def maker(file_list):
    parts = []
    for x in file_list:
        if '.dcm' in x:
            plan = d.read_file(x)
            
            name = str(plan.PatientName)
            temp_dob = str(plan.PatientBirthDate)
            dob_fake = dt.date(int(temp_dob[:4]), int(temp_dob[4:6]), int(temp_dob[6:]))
            dob = dob_fake.strftime("%m/%d/%y")
            date = str(plan.StudyDate)
            part_number = str(plan.PatientID)
            xray_number = 1
            
            parts.append( [name, date, dob, part_number, xray_number] )
            print("Working...")
    
    return parts


def files_in_folder(temp):
    file_list = []
    
    for (dirpath, dirnames, filenames) in os.walk(temp):
        file_list.extend(filenames)
        break
    
    return file_list

def folder_finder(temp):
    folder_list = []
    
    for (dirpath, dirnames, filenames) in os.walk(temp):
        folder_list.extend(dirnames)
        break

    return folder_list

def file_deleter(file_list):
    for file in file_list:
        os.remove(file)
    return


def ecg_count(parts):
    ecg_count = 0
    for participant in parts:
        ecg_count += int(participant[4])
    return ecg_count

	
def main():
    cover_info = []
    temp = 'C:\\Users\\adrian.ridder\\Desktop\\xrays'
    os.chdir(temp)
    for date in folder_finder(temp):
        os.chdir(temp + '\\' + date)                #Change directory to xrays folder
        part_list = maker(files_in_folder((temp + '\\' + date)))
        print(part_list)
        os.chdir(temp)
        sheet_creator(part_list)                  #Makes the spreadsheet using the part_list
        print("Done")
    
if __name__ == '__main__':
    main()

